import openpyxl

wb = openpyxl.load_workbook('Dumentos LaText (1).xlsx')

# Analizar Secciones
ws = wb['Secciones']
print('=== SECCIONES ===')
print('Columnas:', [cell.value for cell in ws[1]])
print('\nPrimeras 5 secciones:')
for i in range(2, min(7, ws.max_row+1)):
    row = [cell.value for cell in ws[i]]
    print(f'Fila {i}: {row}')

# Analizar Bibliografia
print('\n=== BIBLIOGRAFIA ===')
ws_bib = wb['Bibliografia']
print('Columnas:', [cell.value for cell in ws_bib[1]])
print('\nPrimeras 2 referencias:')
for i in range(2, min(4, ws_bib.max_row+1)):
    row = [cell.value for cell in ws_bib[i]]
    print(f'Fila {i}: {row}')
