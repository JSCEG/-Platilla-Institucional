import openpyxl

wb = openpyxl.load_workbook('Dumentos LaText (1).xlsx')

hojas = ['Figuras', 'Tablas', 'Datos Tablas', 'Siglas', 'Glosario', 'Bibliografia']

for nombre_hoja in hojas:
    if nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        print(f'\n=== {nombre_hoja.upper()} ===')
        print('Columnas:', [cell.value for cell in ws[1]])
        
        # Mostrar primeras 2 filas de datos
        if ws.max_row >= 2:
            print('\nPrimeras filas de datos:')
            for i in range(2, min(4, ws.max_row + 1)):
                row = [cell.value for cell in ws[i]]
                print(f'Fila {i}: {row}')
    else:
        print(f'\n⚠️ Hoja "{nombre_hoja}" no encontrada')
